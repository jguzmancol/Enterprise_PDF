import multiprocessing
import os
from queue import Empty

from app.config import CONVERT_TIMEOUT_SECONDS


def get_page_count(pdf_path: str) -> int:
    import fitz

    with fitz.open(pdf_path) as doc:
        return doc.page_count


def pdf_to_docx(pdf_path: str, output_path: str):
    from pdf2docx import Converter

    cv = Converter(pdf_path)
    try:
        cv.convert(output_path, start=0, end=None)
    finally:
        cv.close()


def pdf_to_xlsx(pdf_path: str, output_path: str):
    import pdfplumber
    from openpyxl import Workbook

    wb = Workbook()
    with pdfplumber.open(pdf_path) as pdf:
        first = True
        for i, page in enumerate(pdf.pages):
            sheet_name = f"Page {i + 1}"[:31]
            ws = wb.create_sheet(title=sheet_name) if first or i > 0 else wb.active
            if first:
                ws.title = sheet_name
                first = False

            tables = page.extract_tables()
            if tables:
                for ti, table in enumerate(tables):
                    if ti > 0:
                        ws = wb.create_sheet(title=f"{sheet_name}_t{ti + 1}"[:31])
                    for row_data in table:
                        ws.append(row_data)
            else:
                text = page.extract_text()
                if text:
                    for line in text.strip().split("\n"):
                        ws.append([line])

    if len(wb.sheetnames) == 1 and not any(
        ws.iter_rows(min_row=1, max_row=1, values_only=True) for ws in wb.worksheets
    ):
        ws = wb.active
        ws.append(["No content extracted from PDF"])

    wb.save(output_path)


def _worker(func, pdf_path: str, output_path: str, queue):
    """Process entry point for run_conversion."""
    try:
        func(pdf_path, output_path)
    except Exception as e:
        queue.put(f"{type(e).__name__}: {e}")
    else:
        queue.put(None)


def run_conversion(func, pdf_path: str, output_path: str, timeout: int | None = None) -> str:
    """Run a conversion function in a separate process with a hard timeout.

    Runs in its own process instead of a thread: CPU-bound conversion work
    never starves the event loop via the GIL, and a hung conversion can be
    killed. Blocks the calling thread; call via asyncio.to_thread.
    """
    timeout = timeout or CONVERT_TIMEOUT_SECONDS
    ctx = multiprocessing.get_context("spawn")
    queue = ctx.Queue()
    proc = ctx.Process(target=_worker, args=(func, pdf_path, output_path, queue), daemon=True)
    proc.start()
    try:
        error = queue.get(timeout=timeout)
    except Empty:
        proc.terminate()
        proc.join(timeout=10)
        if proc.is_alive():
            proc.kill()
        raise TimeoutError(f"Conversion timed out after {timeout}s")
    proc.join(timeout=10)
    if proc.exitcode != 0:
        raise RuntimeError(f"Conversion process failed (exit code {proc.exitcode})")
    if error is not None:
        raise RuntimeError(error)
    if not os.path.exists(output_path) or os.path.getsize(output_path) == 0:
        raise RuntimeError("Conversion produced no output")
    return output_path
